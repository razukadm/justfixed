"""Tests for the XLSX export module (exports/xlsx.py).

Written TDD-first: this file exists before xlsx.py. All tests fail with
ImportError until the module is implemented.
"""

from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import openpyxl

from justfixed.domain.investment import Investment
from justfixed.domain.issuer import Issuer, IssuerKind
from justfixed.domain.money import Money
from justfixed.domain.product import ProductType
from justfixed.domain.rates import (
    PostFixedCDI,
    PostFixedCDIPlusSpread,
    PostFixedIPCA,
    Prefixed,
    rate_type_label,
)
from justfixed.engine.conglomerate_report import (
    ConglomerateDetailRow,
    ConglomerateReport,
    ConglomerateSection,
    ConglomerateStatus,
    build_conglomerate_report,
)
from justfixed.engine.fgc import ExposureStatus, fgc_concentration_report_from_projections
from justfixed.engine.projection import ProjectionResult, project
from justfixed.engine.curve import Curve, CurveVertex
from justfixed.exports.xlsx import (
    _FGC_DISPLAY,
    export_conglomerates_xlsx,
    export_curves_xlsx,
    export_investments_xlsx,
)

# ── Shared constants ──────────────────────────────────────────────────────────

ASSUMED_CDI = Decimal("0.12")
PURCHASE = date(2025, 1, 2)
AS_OF = date(2026, 1, 1)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _bank(name: str) -> Issuer:
    return Issuer.create(name, f"{name} S.A.", IssuerKind.COMMERCIAL_BANK)


def _cdb(
    issuer: Issuer,
    principal: str,
    maturity: date,
    purchase: date = PURCHASE,
) -> Investment:
    return Investment.create(
        product=ProductType.CDB,
        issuer=issuer,
        principal=Money.from_reais(principal),
        rate=Prefixed.from_percent("12"),
        purchase_date=purchase,
        maturity_date=maturity,
    )


def _project(inv: Investment) -> ProjectionResult:
    return project(inv, as_of=AS_OF, assumed_cdi=ASSUMED_CDI)


def _status_map(
    cache: list[ProjectionResult] | None,
) -> dict[uuid.UUID, ExposureStatus]:
    """Build the per-investment FGC status map the same way the UI does."""
    if not cache:
        return {}
    report = fgc_concentration_report_from_projections(cache)
    return {
        exp.investment_id: c.current_status
        for c in report.conglomerates
        for exp in c.investments
    }


def _load(data: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(data))


def _ws(data: bytes):
    return _load(data).active


def _date_eq(cell_value: object, expected: date) -> bool:
    """Compare a cell value that may come back as date or datetime."""
    if isinstance(cell_value, datetime):
        return cell_value.date() == expected
    return cell_value == expected


def _make_section(name: str = "TestBank S.A.", n_rows: int = 1) -> ConglomerateSection:
    """Build a ConglomerateSection with n_rows detail rows for testing."""
    principal = Money.from_reais("10000")
    current = Money.from_reais("10500")
    projected = Money.from_reais("11000")
    rows = [
        ConglomerateDetailRow(
            maturity_date=date(2027, i + 1, 1),
            issuer_name="TestBank",
            product=ProductType.CDB,
            principal=principal,
            current_value=current,
            projected_value=projected,
            projected_balance=projected * (n_rows - i),
            fgc_status=ConglomerateStatus.UNDER,
        )
        for i in range(n_rows)
    ]
    return ConglomerateSection(
        conglomerate_name=name,
        investment_count=n_rows,
        total_principal=principal * n_rows,
        total_current_value=current * n_rows,
        total_projected_value=projected * n_rows,
        next_maturity=rows[0].maturity_date,
        summary_fgc_status=ConglomerateStatus.UNDER,
        rows=rows,
    )


# ── Investments: empty / header ───────────────────────────────────────────────

def test_investments_empty_emits_header_only() -> None:
    data = export_investments_xlsx(
        [], projection_cache=None, fgc_status_by_id={}, as_of=AS_OF
    )
    ws = _ws(data)
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value is not None


def test_investments_header_row() -> None:
    data = export_investments_xlsx(
        [], projection_cache=None, fgc_status_by_id={}, as_of=AS_OF
    )
    ws = _ws(data)
    expected = [
        "Emissor", "Conglomerado", "Custodiante", "Produto", "Tipo", "Taxa",
        "Principal", "Vencimento", "Atual", "Projetado", "FGC",
    ]
    actual = [ws.cell(row=1, column=c).value for c in range(1, len(expected) + 1)]
    assert actual == expected


# ── Investments: row values ───────────────────────────────────────────────────

def test_investments_one_row_values() -> None:
    issuer = _bank("Banco Inter")
    inv = _cdb(issuer, "50000", date(2027, 6, 1))
    result = _project(inv)
    status_map = _status_map([result])

    data = export_investments_xlsx(
        [inv], projection_cache=[result], fgc_status_by_id=status_map, as_of=AS_OF
    )
    ws = _ws(data)

    assert ws.cell(row=2, column=1).value == issuer.name
    assert ws.cell(row=2, column=4).value == inv.product.value

    principal_cell = ws.cell(row=2, column=7)
    assert isinstance(principal_cell.value, (int, float))
    assert principal_cell.value == float(inv.principal.amount)

    assert _date_eq(ws.cell(row=2, column=8).value, inv.maturity_date)

    current_cell = ws.cell(row=2, column=9)
    assert isinstance(current_cell.value, (int, float))
    assert current_cell.value == float(result.current_value.amount)

    projected_cell = ws.cell(row=2, column=10)
    assert isinstance(projected_cell.value, (int, float))
    assert projected_cell.value == float(result.gross_at_maturity.amount)

    fgc_cell = ws.cell(row=2, column=11).value
    assert fgc_cell in set(_FGC_DISPLAY.values())


def test_investments_uncached_row_blank_value_cells() -> None:
    inv = _cdb(_bank("Banco Inter"), "50000", date(2027, 6, 1))

    data = export_investments_xlsx(
        [inv], projection_cache=None, fgc_status_by_id={}, as_of=AS_OF
    )
    ws = _ws(data)

    assert ws.cell(row=2, column=9).value is None    # Current blank
    assert ws.cell(row=2, column=10).value is None   # Projected blank
    assert ws.cell(row=2, column=11).value is None   # FGC blank (not in map)


def test_investments_treasury_fgc_cell() -> None:
    inv = Investment.create(
        product=ProductType.TESOURO_PREFIXADO,
        issuer=Issuer.treasury(),
        principal=Money.from_reais("10000"),
        rate=Prefixed.from_percent("12"),
        purchase_date=PURCHASE,
        maturity_date=date(2027, 1, 2),
    )
    result = project(inv, as_of=AS_OF, assumed_cdi=ASSUMED_CDI)
    # Treasury is filtered out of the FGC report → empty map.
    # The issuer-kind branch in the export must still return "N/A".
    status_map = _status_map([result])

    data = export_investments_xlsx(
        [inv], projection_cache=[result], fgc_status_by_id=status_map, as_of=AS_OF
    )
    ws = _ws(data)

    assert ws.cell(row=2, column=11).value == "N/A"


def test_investments_money_cells_are_numeric() -> None:
    inv = _cdb(_bank("Banco Alpha"), "30000", date(2027, 3, 1))
    result = _project(inv)

    data = export_investments_xlsx(
        [inv],
        projection_cache=[result],
        fgc_status_by_id=_status_map([result]),
        as_of=AS_OF,
    )
    ws = _ws(data)

    for col in (7, 9, 10):  # Principal, Current, Projected
        val = ws.cell(row=2, column=col).value
        assert isinstance(val, (int, float)), (
            f"column {col} expected numeric, got {type(val)}"
        )


def test_investments_rows_match_input_order() -> None:
    bank_a = _bank("Alfa")
    bank_b = _bank("Beta")
    bank_c = _bank("Gama")
    inv_a = _cdb(bank_a, "10000", date(2027, 2, 1))
    inv_b = _cdb(bank_b, "20000", date(2028, 3, 1))
    inv_c = _cdb(bank_c, "30000", date(2029, 4, 1))

    data = export_investments_xlsx(
        [inv_a, inv_b, inv_c],
        projection_cache=None,
        fgc_status_by_id={},
        as_of=AS_OF,
    )
    ws = _ws(data)

    assert ws.cell(row=2, column=1).value == bank_a.name
    assert ws.cell(row=3, column=1).value == bank_b.name
    assert ws.cell(row=4, column=1).value == bank_c.name


def test_investments_fgc_reflects_conglomerate_not_individual() -> None:
    # Three R$100k CDBs at the same issuer (same conglomerate).
    # Each individual current value at as_of ≈ R$112k → well under R$250k alone.
    # Conglomerate aggregate ≈ R$336k → OVER R$250k.
    # The export must reflect the conglomerate classification ("over"), not
    # each investment in isolation ("under"). This pins the bug where the old
    # _fgc_value helper checked each investment's own gross_at_maturity against
    # the R$250k limit, missing the per-conglomerate summation entirely.
    issuer = _bank("Banco Inter")
    inv_a = _cdb(issuer, "100000", date(2027, 6, 1))
    inv_b = _cdb(issuer, "100000", date(2027, 9, 1))
    inv_c = _cdb(issuer, "100000", date(2028, 1, 2))
    investments = [inv_a, inv_b, inv_c]
    cache = [_project(inv) for inv in investments]
    status_map = _status_map(cache)

    data = export_investments_xlsx(
        investments,
        projection_cache=cache,
        fgc_status_by_id=status_map,
        as_of=AS_OF,
    )
    ws = _ws(data)

    for row in (2, 3, 4):
        assert ws.cell(row=row, column=11).value == "ACIMA", (
            f"row {row}: expected 'ACIMA' (conglomerate aggregate ≈ R$336k), "
            f"got {ws.cell(row=row, column=11).value!r}"
        )


def test_investments_custodian_column() -> None:
    issuer = _bank("Banco Inter")
    inv_with = Investment.create(
        product=ProductType.CDB,
        issuer=issuer,
        principal=Money.from_reais("10000"),
        rate=Prefixed.from_percent("12"),
        purchase_date=PURCHASE,
        maturity_date=date(2027, 6, 1),
        custodian="XP Investimentos",
    )
    inv_without = Investment.create(
        product=ProductType.CDB,
        issuer=issuer,
        principal=Money.from_reais("10000"),
        rate=Prefixed.from_percent("12"),
        purchase_date=PURCHASE,
        maturity_date=date(2027, 9, 1),
    )

    data = export_investments_xlsx(
        [inv_with, inv_without],
        projection_cache=None,
        fgc_status_by_id={},
        as_of=AS_OF,
    )
    ws = _ws(data)

    assert ws.cell(row=2, column=3).value == "XP Investimentos"
    assert ws.cell(row=3, column=3).value is None


# ── Conglomerates: empty / header ─────────────────────────────────────────────

def test_conglomerates_empty_emits_header_only() -> None:
    report = ConglomerateReport(sections=[], as_of=AS_OF)
    data = export_conglomerates_xlsx(report)
    ws = _ws(data)
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value is not None


def test_conglomerates_header_row() -> None:
    report = ConglomerateReport(sections=[], as_of=AS_OF)
    data = export_conglomerates_xlsx(report)
    ws = _ws(data)
    expected = [
        "Conglomerado", "Investimentos", "Principal", "Atual", "Projetado",
        "Próx. vencimento", "FGC",
    ]
    actual = [ws.cell(row=1, column=c).value for c in range(1, len(expected) + 1)]
    assert actual == expected


# ── Conglomerates: row values ─────────────────────────────────────────────────

def test_conglomerates_one_section_values() -> None:
    inv = _cdb(_bank("Banco Inter"), "50000", date(2027, 6, 1))
    report = build_conglomerate_report([inv], as_of=AS_OF, assumed_cdi=ASSUMED_CDI)
    section = report.sections[0]

    data = export_conglomerates_xlsx(report)
    ws = _ws(data)

    assert ws.cell(row=2, column=1).value == section.conglomerate_name

    inv_count = ws.cell(row=2, column=2).value
    assert isinstance(inv_count, (int, float))
    assert inv_count == section.investment_count

    principal_cell = ws.cell(row=2, column=3)
    assert isinstance(principal_cell.value, (int, float))
    assert principal_cell.value == float(section.total_principal.amount)

    current_cell = ws.cell(row=2, column=4)
    assert isinstance(current_cell.value, (int, float))
    assert current_cell.value == float(section.total_current_value.amount)

    projected_cell = ws.cell(row=2, column=5)
    assert isinstance(projected_cell.value, (int, float))
    assert projected_cell.value == float(section.total_projected_value.amount)

    assert _date_eq(ws.cell(row=2, column=6).value, section.next_maturity)
    assert ws.cell(row=2, column=7).value == _FGC_DISPLAY[section.summary_fgc_status.value]


def test_conglomerates_summary_only_no_detail_rows() -> None:
    # Section has 3 detail rows; the export must emit only 1 summary row.
    section = _make_section(n_rows=3)
    report = ConglomerateReport(sections=[section], as_of=AS_OF)

    data = export_conglomerates_xlsx(report)
    ws = _ws(data)

    assert ws.max_row == 2   # 1 header + 1 summary; detail rows not exported


def test_conglomerates_one_row_per_section() -> None:
    def _section(n: int) -> ConglomerateSection:
        principal = Money.from_reais("10000")
        current = Money.from_reais("10500")
        projected = Money.from_reais("11000")
        rows = [
            ConglomerateDetailRow(
                maturity_date=date(2027, 1, n),
                issuer_name=f"Bank{n}",
                product=ProductType.CDB,
                principal=principal,
                current_value=current,
                projected_value=projected,
                projected_balance=projected,
                fgc_status=ConglomerateStatus.UNDER,
            )
        ]
        return ConglomerateSection(
            conglomerate_name=f"Conglomerate {n}",
            investment_count=1,
            total_principal=principal,
            total_current_value=current,
            total_projected_value=projected,
            next_maturity=rows[0].maturity_date,
            summary_fgc_status=ConglomerateStatus.UNDER,
            rows=rows,
        )

    report = ConglomerateReport(
        sections=[_section(1), _section(2), _section(3)],
        as_of=AS_OF,
    )
    data = export_conglomerates_xlsx(report)
    ws = _ws(data)

    assert ws.max_row == 4   # 1 header + 3 section rows


# ── Validity ──────────────────────────────────────────────────────────────────

def test_both_exports_load_as_valid_xlsx() -> None:
    inv = _cdb(_bank("Banco Inter"), "50000", date(2027, 6, 1))
    result = _project(inv)

    investments_data = export_investments_xlsx(
        [inv],
        projection_cache=[result],
        fgc_status_by_id=_status_map([result]),
        as_of=AS_OF,
    )
    openpyxl.load_workbook(BytesIO(investments_data))

    report = build_conglomerate_report([inv], as_of=AS_OF, assumed_cdi=ASSUMED_CDI)
    cong_data = export_conglomerates_xlsx(report)
    openpyxl.load_workbook(BytesIO(cong_data))


# ── Type column: pt-BR labels ─────────────────────────────────────────────────

def test_investments_type_column_ptbr_labels() -> None:
    """Type column must emit pt-BR labels for all four rate kinds, not English."""
    issuer = _bank("Banco Inter")
    maturity = date(2027, 6, 1)

    rates = [
        Prefixed.from_percent("12"),
        PostFixedCDI.from_percent("112"),
        PostFixedCDIPlusSpread.from_percent("2.05"),
        PostFixedIPCA.from_percent("5.5"),
    ]
    investments = [
        Investment.create(
            product=ProductType.CDB,
            issuer=issuer,
            principal=Money.from_reais("10000"),
            rate=r,
            purchase_date=PURCHASE,
            maturity_date=maturity,
        )
        for r in rates
    ]

    data = export_investments_xlsx(
        investments, projection_cache=None, fgc_status_by_id={}, as_of=AS_OF
    )
    ws = _ws(data)

    # Each Type cell must match rate_type_label (canonical source of truth).
    for row_idx, rate in enumerate(rates, start=2):
        cell = ws.cell(row=row_idx, column=5).value
        assert cell == rate_type_label(rate), (
            f"row {row_idx}: expected {rate_type_label(rate)!r}, got {cell!r}"
        )

    # Explicit pt-BR label assertions — guards against accidentally restoring
    # the old English labels ("Prefixed", "CDI%", "CDI+").
    assert ws.cell(row=2, column=5).value == "Pré"
    assert ws.cell(row=3, column=5).value == "Pós"
    assert ws.cell(row=4, column=5).value == "Pós+"
    assert ws.cell(row=5, column=5).value == "IPCA+"


# ── Curve export ──────────────────────────────────────────────────────────────

_ANCHOR = date(2026, 1, 2)


def _curve(*pairs: tuple[int, str]) -> Curve:
    return Curve(
        anchor=_ANCHOR,
        vertices=tuple(
            CurveVertex(business_days=bd, rate=Decimal(r))
            for bd, r in pairs
        ),
    )


def test_curves_returns_bytes() -> None:
    data = export_curves_xlsx(None, None, None)
    assert isinstance(data, bytes)


def test_curves_reopens_as_valid_xlsx() -> None:
    data = export_curves_xlsx(None, None, None)
    openpyxl.load_workbook(BytesIO(data))


def test_curves_three_sheets_in_order_when_all_none() -> None:
    data = export_curves_xlsx(None, None, None)
    assert _load(data).sheetnames == ["CDI", "PRE", "IPCA"]


def test_curves_three_sheets_in_order_when_all_populated() -> None:
    cdi = _curve((63, "0.120"), (252, "0.130"))
    pre = _curve((63, "0.110"))
    ipca = _curve((252, "0.060"))
    data = export_curves_xlsx(cdi, pre, ipca)
    assert _load(data).sheetnames == ["CDI", "PRE", "IPCA"]


def test_curves_header_row_on_all_sheets() -> None:
    data = export_curves_xlsx(None, None, None)
    wb = _load(data)
    for title in ("CDI", "PRE", "IPCA"):
        ws = wb[title]
        assert ws.cell(row=1, column=1).value == "Dias úteis"
        assert ws.cell(row=1, column=2).value == "Taxa"


def test_curves_vertex_rows_business_days_and_rate() -> None:
    cdi = _curve((63, "0.120"), (126, "0.125"), (252, "0.130"))
    data = export_curves_xlsx(cdi, None, None)
    ws = _load(data)["CDI"]
    assert ws.max_row == 4  # header + 3 vertices
    assert ws.cell(row=2, column=1).value == 63
    assert ws.cell(row=2, column=2).value == float(Decimal("0.120"))
    assert ws.cell(row=3, column=1).value == 126
    assert ws.cell(row=3, column=2).value == float(Decimal("0.125"))
    assert ws.cell(row=4, column=1).value == 252
    assert ws.cell(row=4, column=2).value == float(Decimal("0.130"))


def test_curves_business_days_cell_is_int() -> None:
    cdi = _curve((252, "0.144"))
    val = _load(export_curves_xlsx(cdi, None, None))["CDI"].cell(row=2, column=1).value
    assert isinstance(val, int)


def test_curves_rate_cell_is_float_not_string() -> None:
    cdi = _curve((252, "0.144"))
    val = _load(export_curves_xlsx(cdi, None, None))["CDI"].cell(row=2, column=2).value
    assert isinstance(val, float)
    assert val == float(Decimal("0.144"))


def test_curves_vertex_order_preserved() -> None:
    cdi = _curve((21, "0.110"), (63, "0.120"), (252, "0.130"))
    ws = _load(export_curves_xlsx(cdi, None, None))["CDI"]
    assert ws.cell(row=2, column=1).value == 21
    assert ws.cell(row=3, column=1).value == 63
    assert ws.cell(row=4, column=1).value == 252


def test_curves_none_curve_produces_marker_row() -> None:
    data = export_curves_xlsx(None, None, None)
    wb = _load(data)
    for title in ("CDI", "PRE", "IPCA"):
        ws = wb[title]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "(nenhuma curva carregada)"
        assert ws.cell(row=2, column=2).value is None


def test_curves_empty_vertices_produces_marker_row() -> None:
    empty = Curve(anchor=_ANCHOR, vertices=())
    data = export_curves_xlsx(empty, None, None)
    ws = _load(data)["CDI"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "(nenhuma curva carregada)"
    assert ws.cell(row=2, column=2).value is None


def test_curves_mixed_none_and_populated() -> None:
    cdi = _curve((252, "0.144"))
    data = export_curves_xlsx(cdi, None, None)
    wb = _load(data)
    assert wb["CDI"].max_row == 2  # header + 1 vertex
    assert wb["PRE"].max_row == 2  # header + marker
    assert wb["IPCA"].max_row == 2  # header + marker
    assert wb["CDI"].cell(row=2, column=2).value == float(Decimal("0.144"))
    assert wb["PRE"].cell(row=2, column=1).value == "(nenhuma curva carregada)"
    assert wb["IPCA"].cell(row=2, column=1).value == "(nenhuma curva carregada)"
