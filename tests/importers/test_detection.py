"""Tests for broker detection and unified statement dispatch."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from justfixed.importers.detection import Broker, detect_broker, load_statement
from justfixed.importers.loader_types import LoadResult
from justfixed.persistence.database import Base, make_engine, make_session_factory


FIXTURES   = Path(__file__).parent / "fixtures"
XP_FIXTURE  = FIXTURES / "synthetic_xp_statement.xlsx"
BTG_FIXTURE = FIXTURES / "synthetic_btg_statement.xlsx"
BB_FIXTURE  = FIXTURES / "synthetic_bb_statement.txt"


# ---------- Fixtures ----------


@pytest.fixture
def factory():
    """Fresh in-memory SQLite database per test, with schema created."""
    engine = make_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    yield make_session_factory(engine)
    engine.dispose()


# ---------- detect_broker ----------


class TestDetectBroker:
    def test_xp_fixture_detected_as_xp(self) -> None:
        assert detect_broker(XP_FIXTURE) == Broker.XP

    def test_btg_fixture_detected_as_btg(self) -> None:
        assert detect_broker(BTG_FIXTURE) == Broker.BTG

    def test_unrecognized_file_raises_value_error(self, tmp_path: Path) -> None:
        unknown = tmp_path / "unknown_statement.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(unknown)

        with pytest.raises(ValueError, match="unknown_statement.xlsx"):
            detect_broker(unknown)

    def test_unrecognized_error_message_names_sheets(self, tmp_path: Path) -> None:
        unknown = tmp_path / "mystery.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(unknown)

        with pytest.raises(ValueError, match="Sheet1"):
            detect_broker(unknown)

    def test_nonexistent_path_raises_file_not_found(self, tmp_path: Path) -> None:
        missing = tmp_path / "does_not_exist.xlsx"
        with pytest.raises(FileNotFoundError):
            detect_broker(missing)

    def test_ambiguous_file_raises_value_error(self, tmp_path: Path) -> None:
        ambiguous = tmp_path / "ambiguous.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sua carteira"
        wb.create_sheet("Renda Fixa")
        wb.save(ambiguous)

        with pytest.raises(ValueError, match="Ambiguous"):
            detect_broker(ambiguous)


# ---------- load_statement ----------


class TestLoadStatement:
    def test_xp_fixture_returns_xp_broker_and_load_result(
        self, factory
    ) -> None:
        broker, result = load_statement(XP_FIXTURE, factory)

        assert broker == Broker.XP
        assert isinstance(result, LoadResult)
        assert result.inserted >= 1
        assert result.skipped == 0

    def test_btg_fixture_returns_btg_broker_and_load_result(
        self, factory
    ) -> None:
        broker, result = load_statement(BTG_FIXTURE, factory)

        assert broker == Broker.BTG
        assert isinstance(result, LoadResult)
        assert result.inserted == 2
        assert result.skipped == 0

    def test_xp_load_statement_is_idempotent(self, factory) -> None:
        load_statement(XP_FIXTURE, factory)
        broker, result = load_statement(XP_FIXTURE, factory)

        assert broker == Broker.XP
        assert result.inserted == 0
        assert result.skipped >= 1

    def test_btg_load_statement_is_idempotent(self, factory) -> None:
        load_statement(BTG_FIXTURE, factory)
        broker, result = load_statement(BTG_FIXTURE, factory)

        assert broker == Broker.BTG
        assert result.inserted == 0
        assert result.skipped == 2


# ---------- LoadResult default field (regression guard) ----------


class TestLoadResultDefault:
    def test_skipped_matured_defaults_to_zero(self) -> None:
        """XP and BTG construction sites pass 4 positional args — must not break."""
        result = LoadResult(inserted=10, skipped=2, issuers_created=1, issuers_reused=3)
        assert result.skipped_matured == 0

    def test_skipped_matured_can_be_set(self) -> None:
        result = LoadResult(inserted=5, skipped=4, issuers_created=1, issuers_reused=0, skipped_matured=4)
        assert result.skipped_matured == 4


# ---------- detect_broker — BB ----------


class TestDetectBrokerBB:
    def test_bb_fixture_detected_as_bb(self) -> None:
        assert detect_broker(BB_FIXTURE) == Broker.BB

    def test_txt_without_sisbb_header_raises(self, tmp_path: Path) -> None:
        plain_txt = tmp_path / "random.txt"
        plain_txt.write_text("Just some random text\nNo SISBB header here.\n", encoding="utf-8")
        with pytest.raises(ValueError, match="SISBB"):
            detect_broker(plain_txt)

    def test_unknown_extension_raises(self, tmp_path: Path) -> None:
        bad_file = tmp_path / "statement.csv"
        bad_file.write_text("col1,col2\n1,2\n", encoding="utf-8")
        with pytest.raises(ValueError, match=r"\.csv"):
            detect_broker(bad_file)

    # --- Regression: existing XLSX detection unchanged ---

    def test_xp_fixture_still_detected_as_xp(self) -> None:
        assert detect_broker(XP_FIXTURE) == Broker.XP

    def test_btg_fixture_still_detected_as_btg(self) -> None:
        assert detect_broker(BTG_FIXTURE) == Broker.BTG


# ---------- load_statement — BB ----------


class TestLoadStatementBB:
    def test_bb_fixture_returns_bb_broker(self, factory) -> None:
        broker, result = load_statement(BB_FIXTURE, factory)
        assert broker == Broker.BB

    def test_bb_fixture_returns_load_result(self, factory) -> None:
        _, result = load_statement(BB_FIXTURE, factory)
        assert isinstance(result, LoadResult)

    def test_bb_fixture_inserts_active_rows(self, factory) -> None:
        _, result = load_statement(BB_FIXTURE, factory)
        assert result.inserted == 5

    def test_bb_load_statement_is_idempotent(self, factory) -> None:
        load_statement(BB_FIXTURE, factory)
        broker, result = load_statement(BB_FIXTURE, factory)
        assert broker == Broker.BB
        assert result.inserted == 0
        assert result.skipped == 9
