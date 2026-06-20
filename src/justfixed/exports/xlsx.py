"""XLSX export: investments list and conglomerate summary workbooks."""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO

import openpyxl

from justfixed.domain.investment import Investment
from justfixed.domain.issuer import IssuerKind
from justfixed.domain.rates import rate_type_label
from justfixed.engine.conglomerate_report import ConglomerateReport
from justfixed.engine.curve import Curve
from justfixed.engine.fgc import ExposureStatus
from justfixed.engine.projection import ProjectionResult


_INVESTMENTS_HEADER = [
    "Emissor", "Conglomerado", "Custodiante", "Produto", "Tipo", "Taxa",
    "Principal", "Vencimento", "Atual", "Projetado", "FGC",
]

_CONGLOMERATES_HEADER = [
    "Conglomerado", "Investimentos", "Principal", "Atual", "Projetado",
    "Próx. vencimento", "FGC",
]

_FGC_DISPLAY: dict[str, str] = {
    "under": "ABAIXO",
    "approaching": "PRÓXIMO",
    "over": "ACIMA",
    "not_fgc": "N/A",
}


def _fgc_display(status_value: str) -> str:
    return _FGC_DISPLAY.get(status_value, status_value)


def _fgc_cell(
    inv: Investment,
    fgc_status_by_id: dict[uuid.UUID, ExposureStatus],
) -> str | None:
    # Treasury: FGC does not apply — domain property, checked before the map.
    if inv.issuer.kind == IssuerKind.TREASURY:
        return _FGC_DISPLAY["not_fgc"]
    status = fgc_status_by_id.get(inv.id)
    if status is None:
        return None
    return _fgc_display(status.value)


def export_investments_xlsx(
    investments: list[Investment],
    *,
    projection_cache: list[ProjectionResult] | None,
    fgc_status_by_id: dict[uuid.UUID, ExposureStatus],
    as_of: date,
) -> bytes:
    """Export a flat investments list to XLSX bytes.

    FGC cell rule:
    1. Treasury issuer → "not_fgc".
    2. Investment id in fgc_status_by_id → that ExposureStatus.value.
    3. Not in map (no projection yet) → blank (None).

    Current and Projected cells are blank when the investment has no entry
    in projection_cache. The caller writes to disk.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_INVESTMENTS_HEADER)

    cache_by_id = {r.investment.id: r for r in projection_cache} if projection_cache else {}

    for inv in investments:
        result = cache_by_id.get(inv.id)
        current   = float(result.current_value.amount)    if result else None
        projected = float(result.gross_at_maturity.amount) if result else None
        ws.append([
            inv.issuer.name,
            inv.issuer.conglomerate,
            inv.custodian,
            inv.product.value,
            rate_type_label(inv.rate),
            inv.rate.to_display(),
            float(inv.principal.amount),
            inv.maturity_date,
            current,
            projected,
            _fgc_cell(inv, fgc_status_by_id),
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_conglomerates_xlsx(report: ConglomerateReport) -> bytes:
    """Export a ConglomerateReport summary to XLSX bytes.

    One row per section (detail rows are not exported). The caller writes
    to disk.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_CONGLOMERATES_HEADER)

    for section in report.sections:
        ws.append([
            section.conglomerate_name,
            section.investment_count,
            float(section.total_principal.amount),
            float(section.total_current_value.amount),
            float(section.total_projected_value.amount),
            section.next_maturity,
            _fgc_display(section.summary_fgc_status.value),
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_CURVES_HEADER = ["Dias úteis", "Taxa"]


def export_curves_xlsx(
    cdi: Curve | None,
    pre: Curve | None,
    ipca: Curve | None,
) -> bytes:
    wb = openpyxl.Workbook()

    for i, (title, curve) in enumerate([("CDI", cdi), ("PRE", pre), ("IPCA", ipca)]):
        if i == 0:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)

        ws.append(_CURVES_HEADER)

        if curve is None or not curve.vertices:
            ws.append(["(nenhuma curva carregada)"])
        else:
            for vertex in curve.vertices:
                ws.append([vertex.business_days, float(vertex.rate)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
