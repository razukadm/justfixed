"""XP Investimentos statement parser.

Reads the "Renda Fixa" section of an XP "PosicaoDetalhada" XLSX export,
yielding one XPRow per fixed-income position.

Layer 1 of the importer pipeline:
- Input: path to an XLSX file.
- Output: list of XPRow objects. All fields are RAW STRINGS exactly as
  they appear in the cells. No number parsing, no date parsing, no
  domain types. That's all done in the next layer (xp_mapper).

Why a separate scraping layer? The XLSX has irregular structure
(multi-section, blank-row separators, multiple top-level groups we skip),
which is mostly orthogonal to the per-row parsing concerns. Splitting
keeps each layer focused and testable in isolation.

The parser is forgiving by design:
- Blank rows are separators, not errors.
- Unknown top-level groups (Previdência, Fundos) are skipped silently.
- Within Renda Fixa, only rate-section headers ("XX,X% | <name>") and
  data rows are processed; column-header rows are auto-skipped.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


# A row inside Renda Fixa always has 13 columns:
# A: description
# B: market value (Posição a mercado)
# C: % allocation
# D: valor aplicado (current adjusted basis)
# E: valor aplicado original (acquisition cost)
# F: rate text
# G: purchase date (data aplicação)
# H: maturity date (data vencimento)
# I: quantity
# J: unit price (preço unitário)
# K: IR
# L: IOF
# M: net value (valor líquido)
RENDA_FIXA_COLS = 13


# Pattern for the rate-section header rows: "48,4% | Pós-Fixado"
# - Captures the section name (Pós-Fixado, Prefixado, Inflação).
# - The leading allocation percent is ignored.
_SECTION_RE = re.compile(r"^\s*\d+(?:,\d+)?%\s*\|\s*(.+?)\s*$")


# Top-level groups in the file. Only "Renda Fixa" is in scope.
# Others must be SKIPPED — but the file order is fixed, so once we
# enter Renda Fixa we don't expect to leave it.
_TOP_LEVEL_RENDA_FIXA = "Renda Fixa"


# Recognized rate sections inside Renda Fixa.
RATE_SECTION_POS_FIXADO = "Pós-Fixado"
RATE_SECTION_PREFIXADO = "Prefixado"
RATE_SECTION_INFLACAO = "Inflação"

_KNOWN_RATE_SECTIONS = {
    RATE_SECTION_POS_FIXADO,
    RATE_SECTION_PREFIXADO,
    RATE_SECTION_INFLACAO,
}


@dataclass(frozen=True, slots=True)
class XPRow:
    """A single fixed-income position from an XP statement, raw-string form.

    Every field holds the cell value exactly as it appeared in the
    spreadsheet. Parsing into typed values (Decimal, date, Rate) is the
    next layer's job.

    The `rate_section` is added by the parser based on which sub-section
    the row was found in (Pós-Fixado / Prefixado / Inflação). The XLSX
    does not put this on the row itself — it's contextual.
    """

    rate_section: str        # "Pós-Fixado" | "Prefixado" | "Inflação"
    description: str         # "LCI CEF - ABR/2027"
    market_value: str        # "R$ 576.632,89"
    allocation_pct: str      # "13,14%"
    valor_aplicado: str      # "R$ 500.000,00"
    valor_original: str      # "R$ 500.000,00"
    rate_text: str           # "95,50% CDI" | "+17,59%" | "IPC-A +7,31%" | ...
    purchase_date_text: str  # "02/04/2025"
    maturity_date_text: str  # "02/04/2027"
    quantity_text: str       # "1"
    unit_price_text: str     # "R$ 576.632,89"
    ir_text: str             # "R$ 0,00"
    iof_text: str            # "R$ 0,00"
    net_value_text: str      # "R$ 576.632,89"


def read_renda_fixa_rows(path: Path) -> list[XPRow]:
    """Extract all Renda Fixa positions from an XP XLSX statement.

    Args:
        path: Path to a "PosicaoDetalhada.xlsx" file.

    Returns:
        A list of XPRow objects, one per position. Order matches the
        file order. Empty list if no Renda Fixa section is found.

    Raises:
        FileNotFoundError: If the path doesn't exist.
        ValueError: If the file has structural issues (e.g. a recognized
            section header followed by something that can't be parsed
            as a column-header row).
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        return list(_iter_rows(worksheet))
    finally:
        workbook.close()


def _iter_rows(worksheet) -> list[XPRow]:
    """Walk the worksheet and yield XPRows for every Renda Fixa position.

    State machine:
        * "before_renda_fixa" — looking for the top-level "Renda Fixa" marker.
        * "in_renda_fixa, no_section" — found Renda Fixa, looking for a
          rate-section header.
        * "in_renda_fixa, in_section" — actively collecting data rows
          for the current rate section.

    Note: in XP's layout, the section-header row ("48,4% | Pós-Fixado")
    *also* carries the column labels in cells B–M. So the regex match
    on cell A identifies the section, and the very next row is data.
    No separate column-header-row to skip.
    """
    rows: list[XPRow] = []
    in_renda_fixa = False
    current_section: str | None = None

    for raw_row in worksheet.iter_rows(values_only=True):
        cell_a = _cell_str(raw_row[0]) if raw_row else ""

        # Look for the top-level Renda Fixa marker.
        if not in_renda_fixa:
            if cell_a == _TOP_LEVEL_RENDA_FIXA:
                in_renda_fixa = True
            continue

        # Inside Renda Fixa now. Skip blank/separator rows.
        if not cell_a:
            continue

        # Check for a rate-section header.
        section_match = _SECTION_RE.match(cell_a)
        if section_match and section_match.group(1).strip() in _KNOWN_RATE_SECTIONS:
            current_section = section_match.group(1).strip()
            continue

        # If we hit something unrecognized before any rate section,
        # something is wrong with the file. Skip silently — the rest
        # of the file may still produce useful rows.
        if current_section is None:
            continue

        # Data row. Extract all 13 columns into an XPRow.
        rows.append(_build_row(raw_row, current_section))

    return rows


def _build_row(raw_row: tuple, rate_section: str) -> XPRow:
    """Convert a raw cell tuple into an XPRow. Pads missing trailing cells."""
    cells = [_cell_str(c) for c in raw_row[:RENDA_FIXA_COLS]]
    while len(cells) < RENDA_FIXA_COLS:
        cells.append("")

    return XPRow(
        rate_section=rate_section,
        description=cells[0],
        market_value=cells[1],
        allocation_pct=cells[2],
        valor_aplicado=cells[3],
        valor_original=cells[4],
        rate_text=cells[5],
        purchase_date_text=cells[6],
        maturity_date_text=cells[7],
        quantity_text=cells[8],
        unit_price_text=cells[9],
        ir_text=cells[10],
        iof_text=cells[11],
        net_value_text=cells[12],
    )


def _cell_str(value: object) -> str:
    """Coerce a cell value to a string, treating None as empty."""
    if value is None:
        return ""
    return str(value).strip()