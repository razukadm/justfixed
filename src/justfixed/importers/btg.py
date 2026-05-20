"""BTG Pactual statement parser.

Reads the "Renda Fixa" sheet of a BTG portfolio XLSX export,
yielding one BTGRow per fixed-income position.

Layer 1 of the importer pipeline:
- Input: path to an XLSX file.
- Output: list of BTGRow objects. All fields are RAW STRINGS exactly as
  they appear in the cells, coerced via str(). No number parsing, no date
  parsing, no domain types. That is all done in the next layer (btg_mapper).

Why string-raw? BTG dates arrive from openpyxl as datetime objects and
numbers as int/float. Coercing to str at this layer keeps the mapper's job
symmetric with XP: it always parses strings. Mapper tests then stay simple
and do not depend on openpyxl's type system.

The BTG report is a multi-sheet workbook. Only the "Renda Fixa" sheet is
in scope. Within that sheet, only the "Posicoes Detalhadas" section is
parsed; the earlier "Posicoes" summary table and the trailing
"Posicao Consolidada Por Emissor" aggregation are ignored.

Each Detalhamento sub-section carries its product type and issuer name in a
header row ("Detalhamento > <product> | <issuer>"), so those values are
contextual — they do not appear on the data rows themselves. The parser
threads them through as BTGRow.product and BTGRow.issuer_name.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


RENDA_FIXA_SHEET_NAME = "Renda Fixa"
DETALHAMENTO_ANCHOR = "Posições Detalhadas"
CONSOLIDADA_ANCHOR = "Posição Consolidada Por Emissor"

# Sub-section header: "Detalhamento > <product> | <issuer>"
# Product may contain spaces; issuer may contain spaces and punctuation.
# Leading/trailing whitespace around the pipe is stripped by the regex.
_SECTION_RE = re.compile(
    r"^Detalhamento > (?P<product>[^|]+?)\s*\|\s*(?P<issuer>.+)$"
)


@dataclass(frozen=True, slots=True)
class BTGRow:
    """A single fixed-income position from a BTG statement, raw-string form.

    Every field holds the cell value as str(cell_value). Dates arrive from
    openpyxl as datetime objects; str() produces "YYYY-MM-DD HH:MM:SS".
    Numbers arrive as int or float; str() produces their Python representation.
    This is intentional — the mapper layer does all typed parsing.

    product and issuer_name come from the sub-section header row, not the
    data row itself. The XLSX does not repeat them per data row.

    ativo is kept for parse symmetry; the mapper may drop it since the same
    information is encoded in product + issuer in a different form.
    """

    product:                 str  # "LCI" — from section header
    issuer_name:             str  # "ASSOCIACAO DE POUPANCA E EMPRESTIMO POUPEX"
    ativo:                   str  # "LCI-25I04325998"
    emissao_date_text:       str  # "2025-09-29 00:00:00"
    vencimento_date_text:    str  # "2027-09-16 00:00:00"
    aquisicao_date_text:     str  # "2025-09-29 00:00:00"
    liquidez:                str  # "Sim"
    dias_carencia:           str  # "185"
    data_inicial_liquidez:   str  # "2026-04-02 00:00:00" or "" if absent
    taxa_compra:             str  # "89,00% do CDI"
    quantidade:              str  # "43"
    preco_compra_text:       str  # "1000"
    valor_compra_text:       str  # "43000"
    preco_text:              str  # "1080"
    saldo_bruto_text:        str  # "46440"
    ir_text:                 str  # "-"
    iof_text:                str  # "-"
    saldo_liquido_text:      str  # "46440"


def read_renda_fixa_rows(path: Path) -> list[BTGRow]:
    """Extract all Renda Fixa positions from a BTG XLSX statement.

    Args:
        path: Path to a BTG portfolio XLSX file.

    Returns:
        A list of BTGRow objects, one per position. Order matches the
        file order. Empty list if no Detalhamento section is found.

    Raises:
        FileNotFoundError: If the path does not exist.
        ValueError: If the workbook has no "Renda Fixa" sheet.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if RENDA_FIXA_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {RENDA_FIXA_SHEET_NAME!r} not found in workbook. "
                f"Available sheets: {workbook.sheetnames}"
            )
        worksheet = workbook[RENDA_FIXA_SHEET_NAME]
        return _iter_rows(worksheet)
    finally:
        workbook.close()


def _iter_rows(worksheet) -> list[BTGRow]:
    """Walk the Renda Fixa worksheet and return one BTGRow per data row.

    In BTG exports col A (index 0) is always empty; all row content begins
    in col B (index 1). Anchors, section headers, "Total", "Ativo", and data
    fields are all in col B onward.

    State machine:
        "before_detalhamento" — skip everything until DETALHAMENTO_ANCHOR
            ("Posicoes Detalhadas") in col B. This silently discards the
            Posicoes summary table that appears earlier in the sheet.
        "looking_for_section" — inside the Detalhamento block, looking for
            a sub-section header ("Detalhamento > <product> | <issuer>")
            in col B. Blank and unknown rows are skipped. CONSOLIDADA_ANCHOR
            ends parsing entirely.
        "in_section" — consuming data rows for the current sub-section.
            "Total" in col B ends the sub-section (back to
            looking_for_section). The column-header row is detected by
            col B == "Ativo" — explicit and unambiguous, preferred over a
            row-position offset. A new sub-section header starts a new one
            without requiring an intervening Total. CONSOLIDADA_ANCHOR ends
            parsing. Blank rows are skipped — a sub-section ends only on
            Total, a new header, or Consolidada, never on a blank row.
    """
    rows: list[BTGRow] = []
    state = "before_detalhamento"
    current_product: str | None = None
    current_issuer: str | None = None

    for raw_row in worksheet.iter_rows(values_only=True):
        if not raw_row:
            continue

        col_b = _cell_str(raw_row[1]) if len(raw_row) > 1 else ""

        if state == "before_detalhamento":
            if col_b == DETALHAMENTO_ANCHOR:
                state = "looking_for_section"
            continue

        # CONSOLIDADA_ANCHOR terminates all Detalhamento processing.
        if col_b == CONSOLIDADA_ANCHOR:
            break

        if state == "looking_for_section":
            m = _SECTION_RE.match(col_b)
            if m:
                current_product = m.group("product").strip()
                current_issuer = m.group("issuer").strip()
                state = "in_section"
            # Blank rows and unrecognised rows: stay in looking_for_section.
            continue

        # state == "in_section"

        # A new sub-section header may appear without a preceding Total.
        m = _SECTION_RE.match(col_b)
        if m:
            current_product = m.group("product").strip()
            current_issuer = m.group("issuer").strip()
            continue

        if col_b == "Total":
            state = "looking_for_section"
            continue

        # Column-header row — matched on col B == "Ativo".
        if col_b == "Ativo":
            continue

        # Blank row — carry no data; skip without terminating the sub-section.
        if not col_b:
            continue

        # Data row.
        if current_product is not None and current_issuer is not None:
            rows.append(_build_row(raw_row, current_product, current_issuer))

    return rows


def _build_row(raw_row: tuple, product: str, issuer: str) -> BTGRow:
    """Convert a raw openpyxl row tuple into a BTGRow.

    Column mapping (0-indexed from the row tuple):
      0  = A  always empty in BTG exports
      1  = B  ativo
      2  = C  emissao
      3  = D  vencimento
      4  = E  aquisicao
      5  = F  liquidez
      6  = G  dias_carencia
      7  = H  data_inicial_liquidez
      8  = I  taxa_compra
      9  = J  quantidade
      10 = K  preco_compra
      11 = L  valor_compra
      12 = M  preco
      13 = N  saldo_bruto
      14 = O  ir
      15 = P  iof
      16 = Q  saldo_liquido
    """
    def c(idx: int) -> str:
        return _cell_str(raw_row[idx]) if idx < len(raw_row) else ""

    return BTGRow(
        product=product,
        issuer_name=issuer,
        ativo=c(1),
        emissao_date_text=c(2),
        vencimento_date_text=c(3),
        aquisicao_date_text=c(4),
        liquidez=c(5),
        dias_carencia=c(6),
        data_inicial_liquidez=c(7),
        taxa_compra=c(8),
        quantidade=c(9),
        preco_compra_text=c(10),
        valor_compra_text=c(11),
        preco_text=c(12),
        saldo_bruto_text=c(13),
        ir_text=c(14),
        iof_text=c(15),
        saldo_liquido_text=c(16),
    )


def _cell_str(value: object) -> str:
    """Coerce a cell value to a string, treating None as empty."""
    if value is None:
        return ""
    return str(value).strip()
